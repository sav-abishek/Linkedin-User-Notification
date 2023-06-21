import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


# Function to log in to LinkedIn
def login_to_linkedin(username, password):
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Run in headless mode (without opening browser window)
    service = Service('path/to/chromedriver')  # Replace with actual path to chromedriver
    driver = webdriver.Chrome(service=service, options=chrome_options)

    driver.get('https://www.linkedin.com/login')
    time.sleep(2)

    driver.find_element(By.ID, 'username').send_keys(username)
    driver.find_element(By.ID, 'password').send_keys(password)
    driver.find_element(By.TAG_NAME, 'button').click()

    time.sleep(3)  # Wait for the page to load after login

    return driver


# Function to get the number of unread messages
def get_unread_messages(driver):
    driver.get('https://www.linkedin.com/messaging/')
    time.sleep(2)

    try:
        unread_messages_count = driver.find_element(By.CLASS_NAME, 'msg-overlay-bubble-header__unread-count').text
        unread_messages_count = int(unread_messages_count)
    except NoSuchElementException:
        unread_messages_count = 0

    return unread_messages_count


# Function to get the number of unread notifications
def get_unread_notifications(driver):
    driver.get('https://www.linkedin.com/notifications/')
    time.sleep(2)

    try:
        unread_notifications_count = driver.find_element(By.CLASS_NAME, 'notifications-tab-icon__count').text
        unread_notifications_count = int(unread_notifications_count)
    except NoSuchElementException:
        unread_notifications_count = 0

    return unread_notifications_count


# Function to save data in Excel
def save_data_to_excel(filename, unread_messages_count, unread_notifications_count):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    row = sheet.max_row + 1

    sheet.cell(row=row, column=1).value = unread_messages_count
    sheet.cell(row=row, column=2).value = unread_notifications_count

    wb.save(filename)


# Function to send email notification
def send_email_notification(smtp_server, smtp_port, sender_email, sender_password, recipient_email, unread_messages_count,
                            unread_notifications_count):
    subject = 'LinkedIn Unread Notifications and Messages'
    body = f'Number of unread Messages: {unread_messages_count}\nNumber of unread Notifications: {unread_notifications_count}'

    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = recipient_email
    message['Subject'] = subject
    message.attach(MIMEText(body, 'plain'))

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)


# Main script

# Configuration
linkedin_credentials = [
    {'username': 'linkedin_user1', 'password': 'password1'},
    {'username': 'linkedin_user2', 'password': 'password2'}
]
recipient_emails = ['recipient1@gmail.com', 'recipient2@gmail.com']
excel_filename = 'linkedin_data.xlsx'
smtp_server = 'smtp.gmail.com'
smtp_port = 587
sender_email = 'sender@gmail.com'
sender_password = 'sender_password'

# Loop through LinkedIn users
for credentials in linkedin_credentials:
    username = credentials['username']
    password = credentials['password']

    driver = login_to_linkedin(username, password)
    unread_messages = get_unread_messages(driver)
    unread_notifications = get_unread_notifications(driver)

    save_data_to_excel(excel_filename, unread_messages, unread_notifications)

    driver.quit()

# Loop through recipient emails
for recipient_email in recipient_emails:
    unread_messages_sum = 0
    unread_notifications_sum = 0

    # Calculate sum of unread messages and notifications for all LinkedIn users
    wb = openpyxl.load_workbook(excel_filename)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        unread_messages_sum += sheet.cell(row=row, column=1).value
        unread_notifications_sum += sheet.cell(row=row, column=2).value

    wb.close()

    # Send email notification
    send_email_notification(smtp_server, smtp_port, sender_email, sender_password, recipient_email, unread_messages_sum,
                            unread_notifications_sum)
